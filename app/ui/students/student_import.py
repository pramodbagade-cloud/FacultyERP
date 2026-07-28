"""
FacultyERP
Student Excel Import
--------------------
Handles bulk import of student records from Excel.
"""

#from pathlib import Path
import openpyxl
from tkinter import filedialog
from tkinter import messagebox
from app.services.student_service import StudentService



class StudentImport:
    """Handles student Excel import."""

    REQUIRED_HEADERS = [
        "* First Name",
        "* Last Name",
        "* Mobile Number",
        "* Email Address"
    ]

    TEMPLATE_HEADERS = [
        "* First Name",
        "Middle Name",
        "* Last Name",
        "* Mobile Number",
        "* Email Address",
        "PRN",
        "Aadhaar Number",
        "Gender",
        "Date of Birth",
        "Parent Name",
        "Parent Mobile",
        "Parent Email",
        "Permanent Address",
        "Local Address",
        "Emergency Contact Name",
        "Emergency Contact Number",
        "ABC ID",
        "Remarks"
    ]

    def __init__(
        self,
        parent,
        department_id,
        course_id,
        semester_id,
        division_id,
        academic_year_id,
        admission_year
    ):
        self.parent = parent
        self.department_id = department_id
        self.course_id = course_id
        self.semester_id = semester_id
        self.division_id = division_id
        self.academic_year_id = academic_year_id
        self.admission_year = admission_year
        self.student_service = StudentService()
        self.workbook = None
        self.sheet = None
        self.file_path = None

    # ==========================================================
    # START IMPORT
    # ==========================================================

    def start(self):

        self.file_path = filedialog.askopenfilename(
            title="Select Student Excel File",
            filetypes=[
                ("Excel Workbook", "*.xlsx")
            ]
        )

        if not self.file_path:
            return

        try:

            self.workbook = openpyxl.load_workbook(
                self.file_path
            )

            self.sheet = self.workbook.active

        except Exception as error:

            messagebox.showerror(
                "FacultyERP",
                f"Unable to open Excel file.\n\n{error}"
            )

            return

        if not self.validate_headers():
            return

        students = self.read_excel_data()

        self.validate_students(
            students
        )

        self.validate_duplicate_data(
            students
        )

        self.validate_database_duplicates(
            students
        )

        (
            valid_students,
            invalid_students
        ) = self.get_validation_result(
            students
        )

        if invalid_students:

            message = (
                f"Total Records : {len(students)}\n"
                f"Valid Records : {len(valid_students)}\n"
                f"Invalid Records : {len(invalid_students)}\n\n"
            )

            for student in invalid_students:

                message += (
                    f"Row {student['excel_row']} : "
                    f"{'; '.join(student['errors'])}\n"
                )

            messagebox.showerror(
                "FacultyERP - Validation Errors",
                message
            )

            return

        imported, failed = self.import_students(
            valid_students
        )

        if failed:

            message = (
                f"Imported : {imported}\n"
                f"Failed : {failed}\n\n"
            )

            for student in valid_students:

                if student["errors"]:

                    message += (
                        f"Row {student['excel_row']} : "
                        f"{'; '.join(student['errors'])}\n"
                    )

            messagebox.showerror(
                "FacultyERP - Import Errors",
                message
            )

            return

        messagebox.showinfo(
            "FacultyERP",
            (
                "Student Import Completed Successfully.\n\n"
                f"Total Records : {len(valid_students)}\n"
                f"Imported      : {imported}\n"
                f"Failed        : {failed}"
            )
        )
    # ==========================================================
    # VALIDATE HEADERS
    # ==========================================================

    def validate_headers(self):

        excel_headers = []

        for column in range(
            1,
            len(self.TEMPLATE_HEADERS) + 1
        ):

            value = self.sheet.cell(
                row=1,
                column=column
            ).value

            if value is None:
                value = ""

            excel_headers.append(
                str(value).strip()
            )

        if excel_headers != self.TEMPLATE_HEADERS:

            messagebox.showerror(
                "FacultyERP",
                "Invalid Student Import Template.\n\nPlease download the latest template from FacultyERP."
            )

            return False

        return True
        # ==========================================================
    # READ EXCEL DATA
    # ==========================================================

    def read_excel_data(self):

        students = []

        for row_number, row in enumerate(
            self.sheet.iter_rows(
                min_row=2,
                values_only=True
            ),
            start=2
        ):

            if row is None:
                continue

            if all(
                value in (None, "")
                for value in row
            ):
                continue

            student = {
                "excel_row": row_number,
                "errors": [],
                "first_name": self.get_value(row, 0),
                "middle_name": self.get_value(row, 1),
                "last_name": self.get_value(row, 2),
                "mobile": self.get_value(row, 3),
                "email": self.get_value(row, 4),
                "prn": self.get_value(row, 5),
                "aadhaar_number": self.get_value(row, 6),
                "gender": self.get_value(row, 7),
                "date_of_birth": self.get_value(row, 8),
                "parent_name": self.get_value(row, 9),
                "parent_mobile": self.get_value(row, 10),
                "parent_email": self.get_value(row, 11),
                "permanent_address": self.get_value(row, 12),
                "local_address": self.get_value(row, 13),
                "emergency_contact_name": self.get_value(row, 14),
                "emergency_contact_number": self.get_value(row, 15),
                "abc_id": self.get_value(row, 16),
                "remarks": self.get_value(row, 17)
            }

            students.append(student)

        return students
        # ==========================================================
    # GET CELL VALUE
    # ==========================================================

    def get_value(self, row, index):

        if index >= len(row):
            return ""

        value = row[index]

        if value is None:
            return ""

        return str(value).strip()
        # ==========================================================
    # VALIDATE STUDENTS
    # ==========================================================


    def validate_students(self, students):

        for student in students:

            errors = student["errors"]

            if not student["first_name"]:

                errors.append(
                    "First Name is required."
                )

            if not student["last_name"]:

                errors.append(
                    "Last Name is required."
                )

            mobile = student["mobile"]

            if not mobile:

                errors.append(
                    "Mobile Number is required."
                )

            elif (
                not mobile.isdigit()
                or len(mobile) != 10
            ):

                errors.append(
                    "Mobile Number must contain exactly 10 digits."
                )

            email = student["email"]

            if not email:

                errors.append(
                    "Email Address is required."
                )

            elif (
                "@" not in email
                or "." not in email
            ):

                errors.append(
                    "Invalid Email Address."
                )

            prn = student["prn"]

            if prn and len(prn) < 6:

                errors.append(
                    "PRN appears to be invalid."
                )

            aadhaar = student["aadhaar_number"]

            if aadhaar:

                if (
                    not aadhaar.isdigit()
                    or len(aadhaar) != 12
                ):

                    errors.append(
                        "Aadhaar Number must contain exactly 12 digits."
                    )
    # ==========================================================
    # CHECK DUPLICATES IN EXCEL
    # ==========================================================
        # ==========================================================
    # CHECK DUPLICATES IN EXCEL
    # ==========================================================

    def validate_duplicate_data(self, students):

        mobile_rows = {}

        email_rows = {}

        prn_rows = {}

        aadhaar_rows = {}

        for student in students:

            mobile = student["mobile"]

            if mobile:

                if mobile in mobile_rows:

                    student["errors"].append(
                        f"Duplicate Mobile Number (already used in Row {mobile_rows[mobile]})."
                    )

                else:

                    mobile_rows[mobile] = student["excel_row"]

            email = student["email"].lower()

            if email:

                if email in email_rows:

                    student["errors"].append(
                        f"Duplicate Email Address (already used in Row {email_rows[email]})."
                    )

                else:

                    email_rows[email] = student["excel_row"]

            prn = student["prn"]

            if prn:

                if prn in prn_rows:

                    student["errors"].append(
                        f"Duplicate PRN (already used in Row {prn_rows[prn]})."
                    )

                else:

                    prn_rows[prn] = student["excel_row"]

            aadhaar = student["aadhaar_number"]

            if aadhaar:

                if aadhaar in aadhaar_rows:

                    student["errors"].append(
                        f"Duplicate Aadhaar Number (already used in Row {aadhaar_rows[aadhaar]})."
                    )

                else:

                    aadhaar_rows[aadhaar] = student["excel_row"]

    # ==========================================================
    # CHECK DATABASE DUPLICATES
    # ==========================================================

    def validate_database_duplicates(
            self,
            students
    ):

        existing = StudentService.get_existing_import_data()

        for student in students:

            if student["mobile"]:

                if student["mobile"] in existing["mobiles"]:

                    student["errors"].append(
                        "Mobile Number already exists in FacultyERP."
                    )

            if student["email"]:

                if student["email"].lower() in existing["emails"]:

                    student["errors"].append(
                        "Email Address already exists in FacultyERP."
                    )

            if student["prn"]:

                if student["prn"] in existing["prns"]:

                    student["errors"].append(
                        "PRN already exists in FacultyERP."
                    )

            if student["aadhaar_number"]:

                if (
                    student["aadhaar_number"]
                    in existing["aadhaars"]
                ):

                    student["errors"].append(
                        "Aadhaar Number already exists in FacultyERP."
                    )

    # ==========================================================
    # GET VALID / INVALID STUDENTS
    # ==========================================================

    def get_validation_result(
            self,
            students
    ):

        valid_students = []

        invalid_students = []

        for student in students:

            if student["errors"]:

                invalid_students.append(
                    student
                )

            else:

                valid_students.append(
                    student
                )

        return (
            valid_students,
            invalid_students
        )

    # ==========================================================
    # IMPORT STUDENTS
    # ==========================================================

    def import_students(
            self,
            students
    ):

        imported = 0

        failed = 0

        for student in students:

            success, message = StudentService.import_student(

                student_data=student,

                admission_year=self.admission_year,

                academic_year_id=self.academic_year_id,

                department_id=self.department_id,

                course_id=self.course_id,

                semester_id=self.semester_id,

                division_id=self.division_id

            )

            if success:

                imported += 1

            else:

                failed += 1

                student["errors"].append(message)

        return imported, failed