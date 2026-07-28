"""
FacultyERP
Student Master Window
---------------------
"""

import customtkinter as ctk
import openpyxl


from tkinter import ttk
from tkinter import filedialog
from tkinter import messagebox


from openpyxl.styles import Font
from openpyxl.styles import PatternFill
from openpyxl.styles import Alignment
from openpyxl.styles import Border
from openpyxl.styles import Side

from app.services.student_service import StudentService
from app.services.department_service import DepartmentService
from app.services.course_service import CourseService
from app.services.semester_service import SemesterService
from app.services.division_service import DivisionService
from app.services.academic_year_service import AcademicYearService
from app.ui.students.student_form_window import StudentFormWindow
from app.ui.students.student_import import StudentImport


class StudentWindow:
    """Student Master Window."""

    # ==========================================================
    # INITIALIZE
    # ==========================================================

    def __init__(self, parent):

        self.parent = parent

        self.title_font = ("Segoe UI", 24, "bold")
        self.heading_font = ("Segoe UI", 16, "bold")
        self.normal_font = ("Segoe UI", 12)

        self.container = None
        self.header = None
        self.info_frame = None
        self.toolbar = None
        self.grid_frame = None

        self.student_grid = None

        self.cmb_department = None
        self.cmb_course = None
        self.cmb_semester = None
        self.cmb_division = None
        self.cmb_academic_year = None
        self.cmb_admission_year = None

        self.search_entry = None

        self.department_map = {}
        self.course_map = {}
        self.semester_map = {}
        self.division_map = {}
        self.academic_year_map = {}

        self.selected_department_id = None
        self.selected_course_id = None
        self.selected_semester_id = None
        self.selected_division_id = None
        self.selected_academic_year_id = None

        self.import_students = []

        self.build_ui()

    # ==========================================================
    # BUILD UI
    # ==========================================================

    def build_ui(self):

        for widget in self.parent.winfo_children():
            widget.destroy()

        self.container = ctk.CTkFrame(self.parent)

        self.container.pack(
            fill="both",
            expand=True,
            padx=20,
            pady=20
        )

        self.build_header()

        self.build_academic_information()

        self.build_toolbar()
        self.build_search_bar()
        self.build_student_grid()

        self.load_combo_data()

        self.load_students()
    # ==========================================================
    # HEADER
    # ==========================================================

    def build_header(self):

        self.header = ctk.CTkFrame(self.container)

        self.header.pack(
            fill="x",
            padx=10,
            pady=(0,15)
        )

        ctk.CTkLabel(
            self.header,
            text="Student Master",
            font=self.title_font
        ).pack(
            side="left",
            padx=15,
            pady=15
        )
    
    # ==========================================================
    # ACADEMIC INFORMATION
    # ==========================================================

    def build_academic_information(self):

        self.info_frame = ctk.CTkFrame(self.container)

        self.info_frame.pack(
            fill="x",
            padx=10,
            pady=(0,15)
        )

        for column in (1,3,5):
            self.info_frame.grid_columnconfigure(
                column,
                weight=1
            )

        ctk.CTkLabel(
            self.info_frame,
            text="Department"
        ).grid(
            row=0,
            column=0,
            padx=(20,10),
            pady=15,
            sticky="w"
        )

        self.cmb_department = ctk.CTkComboBox(
            self.info_frame,
            values=[],
            width=220,
            command=self.on_department_changed
        )

        self.cmb_department.grid(
            row=0,
            column=1,
            padx=10,
            pady=15,
            sticky="ew"
        )

        ctk.CTkLabel(
            self.info_frame,
            text="Course"
        ).grid(
            row=0,
            column=2,
            padx=(20,10),
            pady=15,
            sticky="w"
        )

        self.cmb_course = ctk.CTkComboBox(
            self.info_frame,
            values=[],
            width=180,
            command=self.on_course_changed
        )

        self.cmb_course.grid(
            row=0,
            column=3,
            padx=10,
            pady=15,
            sticky="ew"
        )

        ctk.CTkLabel(
            self.info_frame,
            text="Semester"
        ).grid(
            row=0,
            column=4,
            padx=(20,10),
            pady=15,
            sticky="w"
        )

        self.cmb_semester = ctk.CTkComboBox(
            self.info_frame,
            values=[],
            width=150,
            command=self.on_semester_changed
        )

        self.cmb_semester.grid(
            row=0,
            column=5,
            padx=10,
            pady=15,
            sticky="ew"
        )

        ctk.CTkLabel(
            self.info_frame,
            text="Division"
        ).grid(
            row=1,
            column=0,
            padx=(20,10),
            pady=(0,20),
            sticky="w"
        )

        self.cmb_division = ctk.CTkComboBox(
            self.info_frame,
            values=[],
            width=150,
            command=self.on_division_changed
        )

        self.cmb_division.grid(
            row=1,
            column=1,
            padx=10,
            pady=(0,20),
            sticky="ew"
        )

        ctk.CTkLabel(
            self.info_frame,
            text="Academic Year"
        ).grid(
            row=1,
            column=2,
            padx=(20,10),
            pady=(0,20),
            sticky="w"
        )

        self.cmb_academic_year = ctk.CTkComboBox(
            self.info_frame,
            values=[],
            width=180,
            command=self.on_academic_year_changed
        )

        self.cmb_academic_year.grid(
            row=1,
            column=3,
            padx=10,
            pady=(0,20),
            sticky="ew"
        )

        ctk.CTkLabel(
            self.info_frame,
            text="Admission Year"
        ).grid(
            row=1,
            column=4,
            padx=(20,10),
            pady=(0,20),
            sticky="w"
        )

        self.cmb_admission_year = ctk.CTkComboBox(
            self.info_frame,
            values=[
                "2026",
                "2025",
                "2024",
                "2023"
            ],
            width=150
        )

        self.cmb_admission_year.grid(
            row=1,
            column=5,
            padx=10,
            pady=(0,20),
            sticky="ew"
        )

    # ==========================================================
    # DEPARTMENT CHANGED
    # ==========================================================

    def on_department_changed(self, department_name):

        self.selected_department_id = self.department_map.get(
            department_name
        )

        courses = [
            course
            for course in CourseService.get_courses()
            if course.department_id == self.selected_department_id
        ]

        self.course_map.clear()

        course_names = []

        for course in courses:

            course_names.append(
                course.course_name
            )

            self.course_map[
                course.course_name
            ] = course.course_id

        self.cmb_course.configure(
            values=course_names
        )

        if course_names:

            self.cmb_course.set(
                course_names[0]
            )

            self.on_course_changed(
                course_names[0]
            )

    # ==========================================================
    # COURSE CHANGED
    # ==========================================================

    def on_course_changed(self, course_name):

        self.selected_course_id = self.course_map.get(
            course_name
        )

        self.load_divisions()

    # ==========================================================
    # SEMESTER CHANGED
    # ==========================================================

    def on_semester_changed(self, semester_name):

        self.selected_semester_id = self.semester_map.get(
            semester_name
        )

        self.load_divisions()

    # ==========================================================
    # DIVISION CHANGED
    # ==========================================================

    def on_division_changed(self, division_name):

        self.selected_division_id = self.division_map.get(
            division_name
        )

    # ==========================================================
    # ACADEMIC YEAR CHANGED
    # ==========================================================

    def on_academic_year_changed(self, academic_year):

        self.selected_academic_year_id = self.academic_year_map.get(
            academic_year
        )

        self.load_divisions()

    # ==========================================================
    # LOAD DIVISIONS
    # ==========================================================

    def load_divisions(self):

        self.division_map.clear()

        if (
            self.selected_course_id is None
            or self.selected_academic_year_id is None
            or self.selected_semester_id is None
        ):
            self.cmb_division.configure(values=[])
            self.cmb_division.set("")
            return

        divisions = DivisionService.get_divisions_by_course_year_semester(
            self.selected_course_id,
            self.selected_academic_year_id,
            self.selected_semester_id
        )

        division_names = []

        for division in divisions:

            division_names.append(
                division.division_name
            )

            self.division_map[
                division.division_name
            ] = division.division_id

        self.cmb_division.configure(
            values=division_names
        )

        if division_names:

            self.cmb_division.set(
                division_names[0]
            )

            self.on_division_changed(
                division_names[0]
            )

        else:

            self.cmb_division.set("")
    # ==========================================================
    # TOOLBAR
    # ==========================================================

    def build_toolbar(self):

        self.toolbar = ctk.CTkFrame(self.container)

        self.toolbar.pack(
            fill="x",
            padx=10,
            pady=(0,15)
        )

        left_toolbar = ctk.CTkFrame(
            self.toolbar,
            fg_color="transparent"
        )

        left_toolbar.pack(
            side="left",
            padx=10,
            pady=10
        )

        self.btn_download = ctk.CTkButton(
            left_toolbar,
            text="Download Template",
            width=170,
            command=self.download_template
        )

        self.btn_download.pack(
            side="left",
            padx=5
        )

        self.btn_import = ctk.CTkButton(
            left_toolbar,
            text="Import Excel",
            width=140,
            command=self.import_excel
        )

        self.btn_import.pack(
            side="left",
            padx=5
        )

        self.btn_export = ctk.CTkButton(
            left_toolbar,
            text="Export Students",
            width=140,
            command=self.export_students
        )

        self.btn_export.pack(
            side="left",
            padx=5
        )

        self.btn_add = ctk.CTkButton(
            left_toolbar,
            text="Add Student",
            width=140,
            command=self.add_student
        )

        self.btn_add.pack(
            side="left",
            padx=5
        )

        self.btn_edit = ctk.CTkButton(
            left_toolbar,
            text="Edit Student",
            width=140,
            command=self.edit_student
        )

        self.btn_edit.pack(
            side="left",
            padx=5
        )

        self.btn_delete = ctk.CTkButton(
            left_toolbar,
            text="Delete Student",
            width=140,
            command=self.delete_student
        )

        self.btn_delete.pack(
            side="left",
            padx=5
        )



    # ==========================================================
    # SEARCH BAR
    # ==========================================================

    def build_search_bar(self):

        self.search_frame = ctk.CTkFrame(self.container)

        self.search_frame.pack(
            fill="x",
            padx=10,
            pady=(0,15)
        )

        ctk.CTkLabel(
            self.search_frame,
            text="Search Student"
        ).pack(
            side="left",
            padx=(15,10),
            pady=10
        )

        self.search_entry = ctk.CTkEntry(
            self.search_frame,
            width=420,
            placeholder_text="Roll No / Name / PRN / Mobile"
        )

        self.search_entry.pack(
            side="left",
            padx=(0,10),
            pady=10
        )

        self.search_entry.bind(
            "<KeyRelease>",
            self.search_students
        )

    # ==========================================================
    # EXPORT STUDENTS
    # ==========================================================

    def export_students(self):

        messagebox.showinfo(
            "FacultyERP",
            "Student export will be implemented in the next phase."
        )

    # ==========================================================
    # DOWNLOAD TEMPLATE
    # ==========================================================

    def download_template(self):

        file_path = filedialog.asksaveasfilename(
            title="Save Student Import Template",
            defaultextension=".xlsx",
            initialfile="Student_Import_Template.xlsx",
            filetypes=[
                ("Excel Workbook", "*.xlsx")
            ]
        )

        if not file_path:
            return

        workbook = openpyxl.Workbook()

        sheet = workbook.active

        sheet.title = "Students"

        headers = [
            "* First Name",
            "Middle Name",
            "* Last Name",
            "* Mobile Number",
            "* Email Address",
            "PRN",
            "Aadhaar Number",
            "Gender",
            "Date of Birth",
            "Parent Name",
            "Parent Mobile",
            "Parent Email",
            "Permanent Address",
            "Local Address",
            "Emergency Contact Name",
            "Emergency Contact Number",
            "ABC ID",
            "Remarks"
        ]

        widths = [
            20,
            20,
            20,
            18,
            30,
            18,
            20,
            12,
            15,
            25,
            18,
            30,
            40,
            40,
            25,
            22,
            22,
            30
        ]

        from openpyxl.styles import PatternFill
        from openpyxl.styles import Alignment
        from openpyxl.styles import Border
        from openpyxl.styles import Side

        header_fill = PatternFill(
            fill_type="solid",
            fgColor="1F4E78"
        )

        header_font = Font(
            bold=True,
            color="FFFFFF"
        )

        thin = Side(
            style="thin",
            color="D9D9D9"
        )

        for column, header in enumerate(headers, start=1):

            cell = sheet.cell(
                row=1,
                column=column
            )

            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(
                horizontal="center",
                vertical="center"
            )
            cell.border = Border(
                left=thin,
                right=thin,
                top=thin,
                bottom=thin
            )

            sheet.column_dimensions[
                cell.column_letter
            ].width = widths[column - 1]

        sheet.freeze_panes = "A2"

        sheet.auto_filter.ref = sheet.dimensions

        instruction_sheet = workbook.create_sheet(
            title="Instructions"
        )

        instruction_sheet["A1"] = "FacultyERP Student Import Instructions"

        instruction_sheet["A1"].font = Font(
            bold=True,
            size=14
        )

        instruction_sheet["A3"] = "1. Fields prefixed with * are mandatory."
        instruction_sheet["A4"] = "2. Mobile Number must be unique."
        instruction_sheet["A5"] = "3. Email Address must be unique."
        instruction_sheet["A6"] = "4. PRN is optional during admission and can be updated later."
        instruction_sheet["A7"] = "5. Aadhaar Number is optional."
        instruction_sheet["A8"] = "6. Department, Course, Semester, Division and Academic Year are taken from the Student Master screen."
        instruction_sheet["A9"] = "7. Do not modify the column headers."

        instruction_sheet.column_dimensions["A"].width = 120

        workbook.save(file_path)

        messagebox.showinfo(
            "FacultyERP",
            "Student Import Template created successfully."
        )
    # ==========================================================
    # STUDENT GRID
    # ==========================================================

    def build_student_grid(self):

        self.grid_frame = ctk.CTkFrame(self.container)

        self.grid_frame.pack(
            fill="both",
            expand=True,
            padx=10,
            pady=(0,10)
        )
        columns = (
            "Roll No",
            "College ID",
            "Student Name",
            "PRN",
            "Mobile",
            "Department",
            "Semester",
            "Division",
            "Status"
        )

        self.student_grid = ttk.Treeview(
            self.grid_frame,
            columns=columns,
            show="headings"
        )

        column_widths = {
            "Roll No": 60,
            "College ID": 100,
            "Student Name": 200,
            "PRN": 120,
            "Mobile": 110,
            "Department": 150,
            "Semester": 75,
            "Division": 65,
            "Status": 70
        }

        for column in columns:

            self.student_grid.heading(
                column,
                text=column
            )

            self.student_grid.column(
                column,
                width=column_widths[column],
                anchor="center"
            )

        vertical_scroll = ttk.Scrollbar(
            self.grid_frame,
            orient="vertical",
            command=self.student_grid.yview
        )

        self.student_grid.configure(
            yscrollcommand=vertical_scroll.set
        )

        self.student_grid.pack(
            side="left",
            fill="both",
            expand=True
        )

        vertical_scroll.pack(
            side="right",
            fill="y"
        )

    # ==========================================================
    # LOAD STUDENTS
    # ==========================================================

    def load_students(self):

        self.refresh_grid()

    # ==========================================================
    # REFRESH GRID
    # ==========================================================
    def refresh_grid(self):

        for item in self.student_grid.get_children():
            self.student_grid.delete(item)

        students = StudentService.get_all_students()

        for student in students:

            full_name = " ".join(
                value
                for value in [
                    student.first_name,
                    student.middle_name,
                    student.last_name
                ]
                if value
            )

            self.student_grid.insert(
                "",
                "end",
                iid=str(student.student_id),
                values=(
                    student.roll_no,
                    student.college_id,
                    full_name,
                    student.prn,
                    student.mobile,
                    getattr(student, "department_name", ""),
                    getattr(student, "semester_name", ""),
                    getattr(student, "division_name", ""),
                    "Active" if student.is_active else "Inactive"
                )
            )
    # ==========================================================
    # SEARCH STUDENTS
    # ==========================================================
    def search_students(self, event=None):

        keyword = self.search_entry.get().strip().lower()

        for item in self.student_grid.get_children():
            self.student_grid.delete(item)

        students = StudentService.get_all_students()

        for student in students:

            full_name = " ".join(
                value
                for value in [
                    student.first_name,
                    student.middle_name,
                    student.last_name
                ]
                if value
            )

            search_text = " ".join(
                [
                    str(student.roll_no),
                    str(student.college_id),
                    full_name,
                    str(student.prn),
                    str(student.mobile)
                ]
            ).lower()

            if keyword in search_text:

                self.student_grid.insert(
                    "",
                    "end",
                    iid=str(student.student_id),
                    values=(
                        student.roll_no,
                        student.college_id,
                        full_name,
                        student.prn,
                        student.mobile,
                        getattr(student, "department_name", ""),
                        getattr(student, "semester_name", ""),
                        getattr(student, "division_name", ""),
                        "Active" if student.is_active else "Inactive"
                    )
                )
    # ==========================================================
    # LOAD COMBO DATA
    # ==========================================================

    def load_combo_data(self):

        self.department_map.clear()
        self.course_map.clear()
        self.semester_map.clear()
        self.division_map.clear()
        self.academic_year_map.clear()

        departments = DepartmentService.get_departments()
        department_names = []

        for department in departments:
            department_names.append(department.department_name)
            self.department_map[department.department_name] = department.department_id

        self.cmb_department.configure(values=department_names)

        if department_names:
            self.cmb_department.set(department_names[0])
            self.on_department_changed(department_names[0])

        semesters = SemesterService.get_semesters()
        semester_names = []

        for semester in semesters:
            semester_names.append(semester.semester_name)
            self.semester_map[semester.semester_name] = semester.semester_id

        self.cmb_semester.configure(values=semester_names)

        if semester_names:
            self.cmb_semester.set(semester_names[0])
            self.on_semester_changed(semester_names[0])

        academic_years = AcademicYearService.get_academic_years()
        academic_year_names = []

        for year in academic_years:
            academic_year_names.append(year.academic_year)
            self.academic_year_map[year.academic_year] = year.academic_year_id

        self.cmb_academic_year.configure(values=academic_year_names)

        if academic_year_names:
            self.cmb_academic_year.set(academic_year_names[0])
            self.on_academic_year_changed(academic_year_names[0])

    # ==========================================================
    # IMPORT EXCEL
    # ==========================================================

    def import_excel(self):

        if self.selected_department_id is None:

            messagebox.showwarning(
                "FacultyERP",
                "Please select Department."
            )

            return

        if self.selected_course_id is None:

            messagebox.showwarning(
                "FacultyERP",
                "Please select Course."
            )

            return

        if self.selected_semester_id is None:

            messagebox.showwarning(
                "FacultyERP",
                "Please select Semester."
            )

            return

        if self.selected_division_id is None:

            messagebox.showwarning(
                "FacultyERP",
                "Please select Division."
            )

            return

        if self.selected_academic_year_id is None:

            messagebox.showwarning(
                "FacultyERP",
                "Please select Academic Year."
            )

            return

        importer = StudentImport(
            parent=self.parent,
            department_id=self.selected_department_id,
            course_id=self.selected_course_id,
            semester_id=self.selected_semester_id,
            division_id=self.selected_division_id,
            academic_year_id=self.selected_academic_year_id,
            admission_year=self.cmb_admission_year.get()
        )

        importer.start()
        self.load_students()

    # ==========================================================
    # ADD STUDENT
    # ==========================================================

    def add_student(self):

        form = StudentFormWindow(
            self.parent
        )

        self.parent.wait_window(
            form.window
        )

        self.refresh_grid()

    # ==========================================================
    # EDIT STUDENT
    # ==========================================================

    def edit_student(self):

        selected = self.student_grid.focus()

        if not selected:

            messagebox.showwarning(
                "FacultyERP",
                "Please select a student."
            )

            return

        student = StudentService.get_student(
            int(selected)
        )

        if student is None:

            messagebox.showerror(
                "FacultyERP",
                "Student record not found."
            )

            return

        form = StudentFormWindow(
            self.parent,
            student.student_id
        )

        self.parent.wait_window(form.window)

        self.refresh_grid()
    # ==========================================================
    # DELETE STUDENT
    # ==========================================================

    def delete_student(self):

        selected = self.student_grid.focus()

        if not selected:

            messagebox.showwarning(
                "FacultyERP",
                "Please select a student."
            )

            return

        answer = messagebox.askyesno(
            "FacultyERP",
            "Are you sure you want to delete this student?"
        )

        if not answer:

            return

        success = StudentService.delete_student(
            int(selected)
        )

        if success:

            messagebox.showinfo(
                "FacultyERP",
                "Student deleted successfully."
            )

            self.refresh_grid()

        else:

            messagebox.showerror(
                "FacultyERP",
                "Unable to delete student."
            )
