"""
FacultyERP
Excel Importer
--------------
"""

from openpyxl import load_workbook
from app.utils.import_result import ImportResult
from app.utils.validation import Validation


class ExcelImporter:

    @staticmethod
    def import_excel(filepath, headers, mandatory_columns=None):
        if mandatory_columns is None:
            mandatory_columns = []
        result = ImportResult()
        workbook = load_workbook(filepath, data_only=True)
        sheet = workbook["Data"] if "Data" in workbook.sheetnames else workbook.active
        if sheet.max_row < 1:
            result.success = False
            result.errors.append({
                "row": 0,
                "message": "Excel sheet is empty."
            })
            return result
        excel_headers = []
        for cell in sheet[1]:
            value = "" if cell.value is None else str(cell.value).strip()
            value = value.replace("* ", "").replace("*", "").strip()
            excel_headers.append(value)
        if excel_headers != headers:
            result.success = False
            result.errors.append({
                "row": 1,
                "message": "Excel header does not match the template."
            })
            return result
        result.total_rows = sheet.max_row - 1
        for row_number in range(2, sheet.max_row + 1):
            values = []
            blank_row = True
            for cell in sheet[row_number]:
                value = "" if cell.value is None else str(cell.value).strip()
                if value != "":
                    blank_row = False
                values.append(value)
            if blank_row:
                continue
            record = {}
            for index, header in enumerate(headers):
                if index < len(values):
                    record[header] = values[index]
                else:
                    record[header] = ""
            row_has_error = False
            for column in mandatory_columns:
                if Validation.is_blank(record[column]):
                    result.add_error(row_number, f"{column} is mandatory.")
                    row_has_error = True
            if row_has_error:
                continue
            result.add_record(record)
        return result