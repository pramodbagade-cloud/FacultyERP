"""
FacultyERP
Excel Template Utility
----------------------
"""

from openpyxl import Workbook
from openpyxl.styles import PatternFill
from openpyxl.styles import Font
from openpyxl.styles import Alignment
from openpyxl.worksheet.table import Table
from openpyxl.worksheet.table import TableStyleInfo
from openpyxl.utils import get_column_letter


class ExcelTemplate:

    @staticmethod
    def create(filepath, headers, mandatory_columns=None, sample_rows=None):
        if mandatory_columns is None:
            mandatory_columns = []
        if sample_rows is None:
            sample_rows = []
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Data"
        instruction_sheet = workbook.create_sheet(title="Instructions")
        instruction_sheet.append(["Rule", "Description"])
        header_fill = PatternFill(fill_type="solid", start_color="1F4E78", end_color="1F4E78")
        header_font = Font(bold=True, color="FFFFFF")
        header_alignment = Alignment(horizontal="center", vertical="center")
        excel_headers = []
        for header in headers:
            if header in mandatory_columns:
                excel_headers.append(f"* {header}")
            else:
                excel_headers.append(header)
        sheet.append(excel_headers)
        for row in sample_rows:
            sheet.append(row)
        for cell in sheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
        last_column = get_column_letter(len(excel_headers))
        last_row = max(2, sheet.max_row)
        table = Table(displayName="ImportTable", ref=f"A1:{last_column}{last_row}")
        style = TableStyleInfo(name="TableStyleMedium2", showFirstColumn=False, showLastColumn=False, showRowStripes=True, showColumnStripes=False)
        table.tableStyleInfo = style
        sheet.add_table(table)
        sheet.freeze_panes = "A2"
        for column_cells in sheet.columns:
            length = max(len(str(cell.value)) if cell.value is not None else 0 for cell in column_cells)
            sheet.column_dimensions[column_cells[0].column_letter].width = length + 5
        instruction_sheet.append(["Mandatory Columns", "Columns marked with * are compulsory."])
        instruction_sheet.append(["Department", "Department name must exactly match the Department Master."])
        instruction_sheet.append(["Designation", "Designation must exactly match the Designation Master."])
        instruction_sheet.append(["Employee Code", "Must be unique."])
        instruction_sheet.append(["Mobile", "Should contain exactly 10 digits."])
        instruction_sheet.append(["Email", "Enter a valid email address."])
        instruction_sheet.append(["Date Format", "DD-MM-YYYY"])
        instruction_sheet.append(["Do Not", "Do not rename column headers."])
        instruction_sheet.append(["Do Not", "Do not delete mandatory columns."])
        for cell in instruction_sheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
        instruction_sheet.freeze_panes = "A2"
        for column_cells in instruction_sheet.columns:
            length = max(len(str(cell.value)) if cell.value is not None else 0 for cell in column_cells)
            instruction_sheet.column_dimensions[column_cells[0].column_letter].width = length + 5
        workbook.save(filepath)